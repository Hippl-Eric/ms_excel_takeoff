import os
import unittest

from openpyxl import load_workbook, Workbook
from dotenv import load_dotenv

import helpers

class TestHelperFunctions(unittest.TestCase):

    def setUp(self):
        self.directory = os.getenv("TEST_DIR")
        self.wb = load_workbook(filename=f"{self.directory}\\test_workbook.xlsx")
        self.ws = self.wb.active
        self.all_cells = tuple(self.ws.rows)
        self.num_col = len(self.all_cells[0])

    def tearDown(self):
        del self.wb
        del self.ws
        del self.all_cells
        del self.num_col

    def test_cell_search(self):
        result_1 = self.ws['A1']
        result_2 = self.ws['B1']
        result_3 = self.ws['C1']
        result_4 = self.ws['C5']
        self.assertEqual(helpers.cell_search(self.all_cells, 'TEXT'), result_1)
        self.assertEqual(helpers.cell_search(self.all_cells, 'more Text'), result_2)
        self.assertEqual(helpers.cell_search(self.all_cells, '=not_a_formula'), result_3)
        self.assertEqual(helpers.cell_search(self.all_cells, 'SOME_TEXT'), result_4)
        with self.assertRaises(AssertionError):
            helpers.cell_search(self.all_cells, "foobar")

    def test_copy_row(self):
        base_row_idx = 3
        num_rows = 5
        for base_row in self.ws.iter_rows(min_row=base_row_idx, max_row=base_row_idx, max_col=self.num_col):
            helpers.copy_row(self.ws, base_row, num_rows)
        self.wb.save(f'{self.directory}\\result_copy_row.xlsx')
       
    def test_correct_formula(self):
        wb = Workbook()
        ws = wb.active
        ws['J5'] = "=J4+K$5+$J$6+J$7+$J8+SUM($K1:K2)+SUM(K8:$K$10)+SUM(L4:L6)"
        
        # Set function paramaters
        cell = ws['J5']
        start_row_idx = 5
        int_count = 2
        helpers.correct_formula(cell, start_row_idx, int_count)
        
        # Expected return value
        val = "=J4+K$5+$J$8+J$9+$J10+SUM($K1:K2)+SUM(K10:$K$12)+SUM(L4:L8)"
        self.assertEqual(cell.value, val)
        
if __name__ == "__main__":
    load_dotenv()
    unittest.main()
    