from openpyxl import load_workbook
from openpyxl.formula.translate import Translator

def main():

    # Set iteration count
    num_rows = 2

    # Load Template File
    wb = load_workbook(filename="input.xlsx")
    ws = wb["Sheet1"]

    # Insert Rows
    ws.insert_rows(2, num_rows)

    # Copy cells from top row to all inserted rows
    for n in range(1, num_rows + 1):
        for row in ws.iter_rows(min_row=1, max_col=20, max_row=1):
            for cell in row:

                # Access the new cell
                new_cell = cell.offset(row=n, column=0)

                # Translate the formula (if present) or just the value
                try:
                    new_cell.value = Translator(cell.value, origin=cell.coordinate).translate_formula(new_cell.coordinate)
                except TypeError:
                    new_cell.value = cell.value

    file_name = "output.xlsx"
    wb.save(file_name)

    # Success
    return print("Success")

if __name__ == "__main__":
    main()