import os

from openpyxl import load_workbook
from dotenv import load_dotenv

from helpers import cell_search, copy_row, correct_formula, correct_comment_height

def create_new_takeoff(wb, project_name, num_rows, drilled):
    """Description

    [wb] openpyxl wb object
    [project_name] string
    [num_rows] non-negative int
    [drilled] bool
    """

    # Load base template worksheet
    ws = wb["Takeoff-SB"]

    # Load cells and determine number of rows and columns
    all_cells = tuple(ws.rows)
    len_rows = len(all_cells)
    len_col = len(all_cells[0])

    # Search for project name, drilled/driven, and first row locations
    name_cell = cell_search(all_cells, "PROJECTNAME -  TAKEOFF")
    drill_cell = cell_search(all_cells, "DRIVEN/DRILLED")
    title_row_cell = cell_search(all_cells, "SB Nos.")

    # Input project name
    name_cell.value = f"{project_name} - Takeoff"

    # Input drilled or driven
    if drilled:
        drill_cell.value = "DRILLED"
    else:
        drill_cell.value = "DRIVEN"
        
    # Correct comment heights
    for row in all_cells:
        for cell in row:
            if cell.comment:
                correct_comment_height(cell)

    # Locate the first cell row
    first_cell_row_index = title_row_cell.row + 1

    # Copy the first row values and styling, and paste on all added rows
    for first_cell_row in ws.iter_rows(min_row = first_cell_row_index, max_row = first_cell_row_index, max_col = len_col):
        copy_row(ws, first_cell_row, num_rows)
        
    # Fix formulas above first cell row
    for row in ws.iter_rows(min_row = 1, max_row = first_cell_row_index - 1, max_col = len_col):
        for cell in row:
            correct_formula(cell=cell, start_row_idx=first_cell_row_index, int_count=num_rows)
    
    # Fix formulas below inserted rows
    for row in ws.iter_rows(min_row = first_cell_row_index + num_rows + 1, max_row = len_rows + num_rows, max_col = len_col):
        for cell in row:
            correct_formula(cell=cell, start_row_idx=first_cell_row_index, int_count=num_rows)

    # Change value for "SB Nos."
    sb_column = title_row_cell.column
    start_sb = 1
    for row_num in range(first_cell_row_index, first_cell_row_index + num_rows + 1):
        ws.cell(row=row_num, column=sb_column, value=str(start_sb))
        start_sb += 1

    # Correct the print area
    old_area = ws.print_area[0]
    colon_idx = old_area.find(":")
    end_coordinate = old_area[colon_idx + 1:]

    # Convert end_coordinate to cell, offset, return to coordinate
    end_coordinate_cell_obj = ws[end_coordinate]
    end_coordinate_new_cell_obj = end_coordinate_cell_obj.offset(row=num_rows, column=0)
    end_coordinate_new = end_coordinate_new_cell_obj.coordinate

    # Set print area
    ws.print_area = f"{old_area[:colon_idx + 1]}{end_coordinate_new}"
