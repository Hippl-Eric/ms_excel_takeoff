import math

from openpyxl.formula.translate import Translator
from openpyxl.formula import Tokenizer
from openpyxl.utils.cell import coordinate_from_string

from copy import copy

def cell_search(cells, value):
    """Return cell location if equal to value
    
    [cells] lump tuple of all cells.
    [value] value of any type to be checked.
    """
    for row in cells:
        for cell in row:
            if cell.value == value:
                return cell
    raise AssertionError(f"value: {value} not found") 

def copy_row(work_sheet, base_row, int_count):
    """Copy all attributes of the [base_row] and paste [int_count] number of rows below the [base_row]
    
    [work_sheet] active worksheet.
    [base_row] must be a tuple containing a single row cell range.
    [int_count] must be a positive integer.
    """
    base_row_idx = base_row[0].row

    # Create list of style attributes to check later
    # https://openpyxl.readthedocs.io/en/stable/api/openpyxl.styles.styleable.html#openpyxl.styles.styleable.StyleableObject
    style_list = ["alignment", "border", "fill", "font", "number_format", "protection", "quotePrefix"]

    # Correct any merged cells below the base row
    correct_merge_cells(work_sheet, base_row_idx, int_count)
    
    # Correct worksheet row heights for rows below the base row
    correct_row_heights(work_sheet, base_row_idx, int_count)
    
    # Correct List drop down menus
    correct_data_validator(work_sheet, base_row_idx, int_count)
    
    # Insert rows below the base_row
    work_sheet.insert_rows(base_row_idx + 1, int_count)
    
    # Match row heights of inserted rows with base row height
    base_row_height = work_sheet.row_dimensions[base_row_idx].height
    if base_row_height:
        for i in range(base_row_idx+1,base_row_idx+1+int_count):
            work_sheet.row_dimensions[i].height = base_row_height
    
    # Iterate over each cell in the base row
    for cell in base_row:
        
        # Correct formulas in base row with coordinate or range greater than base row index prior to copy
        correct_formula(cell=cell, start_row_idx=base_row_idx, int_count=int_count)

        # Copy cell value and styling and paste to all inserted rows
        for n in range(1, int_count + 1):

            # Select the new cell
            new_cell = cell.offset(row=n, column=0)

            # Translate the formula (if present) or just the value
            try:
                new_cell.value = Translator(cell.value, origin=cell.coordinate).translate_formula(new_cell.coordinate)
            except TypeError:
                new_cell.value = cell.value

            # Copy any styling
            if cell.has_style:
                for style in style_list:
                    setattr(new_cell, style, copy(getattr(cell, style)))

def correct_merge_cells(work_sheet, base_row_idx, int_count):
    
    # Create a list of all merged cell ranges BELOW the base row
    merged_cells_list = [item.coord for item in work_sheet.merged_cells.ranges if item.max_row > base_row_idx]
    
    for cell_range in merged_cells_list:
        
        # Un-merge the oringal range
        work_sheet.unmerge_cells(cell_range)
        
        # Parse and merge the new range
        parsed_cell_range = correct_row_index(cell_range, base_row_idx, int_count)
        work_sheet.merge_cells(parsed_cell_range)
        
def correct_row_heights(work_sheet, base_row_idx, int_count):
    
    # Create list of initial row heights below base row
    work_sheet_row_heights = {idx: dim.height for idx, dim in work_sheet.row_dimensions.items() if idx > base_row_idx}
    for idx, height in work_sheet_row_heights.items():
        
        # Delete the original row height object
        del work_sheet.row_dimensions[idx]
        
        # Set new row height
        work_sheet.row_dimensions[idx + int_count].height = height
            
def correct_data_validator(work_sheet, base_row_idx, int_count):
    
    # Currently only checking for and parsing "lists"
    # https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.datavalidation.html?highlight=drop%20down%20list#openpyxl.worksheet.datavalidation.DataValidation.type
    if work_sheet.data_validations.dataValidation:
        for data_valid in work_sheet.data_validations.dataValidation:
            if data_valid.type == "list":
                if data_valid.formula1:
                    data_valid.formula1 = correct_row_index(data_valid.formula1, base_row_idx, int_count)
                if data_valid.formula2:
                    data_valid.formula2 = correct_row_index(data_valid.formula2, base_row_idx, int_count)

def correct_formula(cell, start_row_idx, int_count):
    """
    Correct the [cell]'s formula after [int_count] number of rows inserted below [start_row_idx]
    Mimics Excel's built in functionality for a cell's formula to be relative (not absolute)
    """
    # Parse only cells with formulas
    if cell.data_type == 'f':
        
        # Tokenize formula and make list of all coordinates and ranges
        formula_str = cell.value
        tok = Tokenizer(formula_str)
        original_vals = [t.value for t in tok.items if t.subtype == 'RANGE']
        
        # Create list for parsed values to be added to
        parsed_vals = []
        
        # Check all original coordinate and range values
        for orig_cell_range in original_vals:
            
            # Parse
            parsed_cell_range = correct_row_index(orig_cell_range, start_row_idx, int_count)
                
            # Add parsed values to the parsed list (may be unchanged)
            parsed_vals.append(parsed_cell_range)
            
        # Replace all original coordinates of formula with parsed coordinates
        for orig, parsed in zip(original_vals, parsed_vals):
            formula_str = formula_str.replace(orig, parsed, 1)
        
        # Set the cell's value to new formula string
        cell.value = formula_str
    
def correct_row_index(cell_coordinate, start_row_idx, int_count):
    """
    Helper function to correct_formula()
    Determines whether cell coordinate or range needs to be corrected (row index > [start_row_idx])
    Adds [int_count] to row index and returns coordinate or range
    """
    
    # If range, break into coordinates
    if ":" in cell_coordinate:
        start, end = range_to_coords(cell_coordinate)
        start_parsed = correct_row_index(start, start_row_idx, int_count)
        end_parsed = correct_row_index(end, start_row_idx, int_count)
        return f"{start_parsed}:{end_parsed}"
        
    # Break into column, row
    col, row = coordinate_from_string(cell_coordinate)
    
    # Add int_count to row index
    if row > start_row_idx:
        new_row = row + int_count
        cell_coordinate = cell_coordinate.replace(str(row), str(new_row))
    
    return cell_coordinate

def range_to_coords(range_coord):
    """
    Take a range (A13:B15) and return two coordinates (A13, B15)
    """
    colon_idx = range_coord.find(":")
    first_coord = range_coord[0:colon_idx]
    second_coord = range_coord[colon_idx+1:]
    return first_coord, second_coord

def correct_comment_height(cell):
    num_lines = comment_line_len(cell._comment.text)
    cell._comment.height = num_lines * 11 * 1.85 # Num_lines * font size * 1.85
        
def comment_line_len(comment_str):
    max_line_length = 22 # Average (may not be perfect, depends on char width in pixels)
    
    # Split at newline
    new_line_list = comment_str.splitlines()
    
    # Calculate num of rows based on max_line_length
    count = 0
    for string in new_line_list:
        if len(string) > 0:
            count += math.ceil(len(string) / max_line_length)
        else:
            count += 1
    return count
