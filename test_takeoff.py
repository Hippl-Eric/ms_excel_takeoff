import os
import unittest

from openpyxl import load_workbook
from dotenv import load_dotenv

import takeoff

class TestHelperFunctions(unittest.TestCase):
    def setUp(self):

        # Grab and set inputs
        test_dir = os.getenv("TEST_DIR")
        test_file = f"\\{os.getenv('TEST_FILE')}"
        check_file = f"\\{os.getenv('CHECK_FILE')}"
        num_rows = int(os.getenv("NUM_ROWS"))
        project_name = "Unit_Test"
        if os.getenv("DRILLED") == "True":
            drilled = True
        else:
            drilled = False
        
        # Create a takeoff (excel) file to test
        takeoff.create_new_takeoff(test_file, project_name, num_rows, drilled, test_dir, test_dir)

        # Load the test workbook
        self.test_wb = load_workbook(filename = f"{test_dir}\\Unit_Test\\PRICING\\Takeoff - Unit_Test.xlsx")
        self.test_ws = self.test_wb["Takeoff-SB"]
        self.test_cells = tuple(self.test_ws)

        # Load the check workbook
        self.check_wb = load_workbook(filename = f"{test_dir}//{check_file}")
        self.check_ws = self.check_wb["Takeoff-SB"]
        self.check_cells = tuple(self.check_ws)

    def tearDown(self):
        del self.test_wb
        del self.test_ws
        del self.test_cells
        
        del self.check_wb
        del self.check_ws
        del self.check_cells

    def test_num_rows(self):

        # Check num rows
        self.assertEqual(len(self.test_cells), len(self.check_cells))

    def test_num_cols(self):

        # Check num columns
        self.assertEqual(len(self.test_cells[0]), len(self.check_cells[0]))

    def test_print_area(self):

        # Check print area
        self.assertEqual(self.test_ws.print_area[0], self.check_ws.print_area[0])

    def test_cell_vals(self):

        # Check each cell
        for test_row, check_row in zip(self.test_cells, self.check_cells):
            for test_cell, check_cell in zip(test_row, check_row):

                # Check value, coordinate, and has_style
                self.assertEqual(test_cell.value, check_cell.value, f"Cell: --{test_cell.coordinate}--, values not equal")
                self.assertEqual(test_cell.coordinate, check_cell.coordinate)
                self.assertEqual(test_cell.has_style, check_cell.has_style, f"Cell: --{test_cell.coordinate}--, .has_style not equal")

    def test_correct_merge_cells(self):
        test_ranges = [merge_range.coord for merge_range in self.test_ws.merged_cells.ranges]
        check_ranges = [merge_range.coord for merge_range in self.check_ws.merged_cells.ranges]
        self.assertCountEqual(test_ranges, check_ranges)
        
    def test_correct_row_heights(self):
        test_heights = {key: dim.height for key, dim in self.test_ws.row_dimensions.items()}
        check_heights = {key: dim.height for key, dim in self.check_ws.row_dimensions.items()}
        self.assertDictEqual(test_heights, check_heights)
    
    def test_cell_style(self):

        # Styles to check
        style_list = ["alignment", "border", "fill", "font", "number_format", "protection", "quotePrefix"]

        # Check each cell
        for test_row, check_row in zip(self.test_cells, self.check_cells):
            for test_cell, check_cell in zip(test_row, check_row):

                # Only test cells with style
                if test_cell.has_style or check_cell.has_style:
                    for style in style_list:
                        
                        # Get the style objects which are values for each style attribute, default None if not present
                        test_style_obj = getattr(test_cell, style, None)
                        check_style_obj = getattr(check_cell, style, None)
                        self.comp_dict(dict_1=test_style_obj, dict_2=check_style_obj, traversal=[], style=[style], coordinate=test_cell.coordinate)

                                
    def comp_dict(self, dict_1, dict_2, style=[], traversal=[], coordinate=None):
        """
        Helper function to recursively return each style object's attributes.
        Style objects can contain nested objects.
        Return, error message if attribute does not match
        """
        try:
            
            # Ensure dict_1 and dict_2 are objects or dict types
            # Will return TypeError if not dict type
            var_dict_1 = vars(dict_1)
            var_dict_2 = vars(dict_2)
        except TypeError:
            
            # Not dict type - compare values
            self.assertEqual(dict_1, dict_2, f"Cell: --{coordinate}--, {style}-{traversal} not equal")
        else:
            
            # dict_1 and dict_2 are in fact dicts, get their attributes and check each one
            for style_attr in var_dict_1:
                val_1 = getattr(dict_1, style_attr)
                val_2 = getattr(dict_2, style_attr)
                
                # Use a traversal list to track nested objects for error messaging
                traversal += [style_attr]
                style = style
                coordinate = coordinate
                
                # Recursively check each attribute
                self.comp_dict(val_1, val_2, style=style, traversal=traversal, coordinate=coordinate)
                traversal = [] 
                # TODO Traversal not working perfectly yet, used to follow nested object trail... not really necessary
            
if __name__ == "__main__":
    load_dotenv()
    unittest.main()
