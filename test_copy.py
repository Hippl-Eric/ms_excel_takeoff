from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from copy import copy

def main():

    # Set iteration count
    num_rows = 6

    # Load Template File
    wb = load_workbook(filename="input.xlsx")
    ws = wb["Sheet1"]

    # Insert Rows
    ws.insert_rows(2, num_rows)

    # Create list of style attributes to check later
    # https://openpyxl.readthedocs.io/en/stable/api/openpyxl.styles.styleable.html#openpyxl.styles.styleable.StyleableObject
    style_list = ["alignment", "border", "fill", "font", "number_format", "protection", "quotePrefix"]

    # Copy cells from top row to all inserted rows
    for n in range(1, num_rows + 1):
        for row in ws.iter_rows(min_row=1, max_col=20, max_row=1):
            for cell in row:

                # Select the new cell
                new_cell = cell.offset(row=n, column=0)

                # Translate the formula (if present) or just the value
                try:
                    new_cell.value = Translator(cell.value, origin=cell.coordinate).translate_formula(new_cell.coordinate)
                except TypeError:
                    new_cell.value = cell.value

                # Copy any styling
                if cell.has_style:
                    for style in style_list:
                        setattr(new_cell, style, copy(getattr(cell, style)))

    file_name = "output.xlsx"
    wb.save(file_name)

    # Success
    return print("Success")

if __name__ == "__main__":
    main()